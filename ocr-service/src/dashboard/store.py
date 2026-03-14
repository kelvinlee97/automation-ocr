"""
内存处理记录存储
使用环形缓冲区保存最近 N 条处理记录，支持分页查询和 Excel 历史回填
"""
import os
import threading
import uuid
from collections import deque
from typing import Optional

from openpyxl import load_workbook

from .models import ProcessRecord


# 内存中最多保留的记录数
MAX_RECORDS = 500


class ProcessStore:
    """线程安全的处理记录环形缓冲区"""

    def __init__(self):
        self._records: deque[ProcessRecord] = deque(maxlen=MAX_RECORDS)
        self._lock = threading.Lock()

    def add(self, record: ProcessRecord) -> None:
        """新增记录（最新的在前）"""
        with self._lock:
            self._records.appendleft(record)

    def list(
        self,
        offset: int = 0,
        limit: int = 20,
        phone: Optional[str] = None,
    ) -> tuple[list[ProcessRecord], int]:
        """
        分页查询，支持手机号模糊搜索
        返回 (记录列表, 筛选后总数)
        """
        with self._lock:
            records = list(self._records)

        if phone:
            records = [r for r in records if phone in r.phone]

        total = len(records)
        return records[offset:offset + limit], total

    def get(self, recordId: str) -> Optional[ProcessRecord]:
        """按 ID 查找单条记录"""
        with self._lock:
            for r in self._records:
                if r.id == recordId:
                    return r
        return None

    @property
    def count(self) -> int:
        with self._lock:
            return len(self._records)

    @property
    def lastProcessedAt(self) -> Optional[str]:
        """最近一条记录的时间"""
        with self._lock:
            if self._records:
                return self._records[0].timestamp
        return None


def backfill_from_excel(store: ProcessStore, excel_path: str) -> int:
    """
    从 Excel Sheet2 回填历史记录到内存（不含步骤详情）
    仅在服务启动时调用一次
    返回回填的记录数
    """
    if not os.path.exists(excel_path):
        return 0

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        # 查找 Receipts sheet
        sheet_name = None
        for name in wb.sheetnames:
            if "receipt" in name.lower():
                sheet_name = name
                break

        if not sheet_name:
            return 0

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=2, values_only=True))

        if not rows:
            return 0

        # 按 write_receipt 写入的 12 列顺序：
        # [序号, 提交时间, 手机号码, 身份证号码, 单据号, 识别品牌, 匹配品牌,
        #  消费金额(RM), 是否合格, 不合格原因, OCR置信度, 图片路径]
        count = 0
        for row in rows:
            if not row or not row[0]:
                continue

            # 解析置信度（可能是 "72.00%" 格式）
            confidenceRaw = str(row[10]) if len(row) > 10 and row[10] else "0"
            confidence = float(confidenceRaw.replace("%", "")) / 100.0 if "%" in confidenceRaw else 0.0

            # 解析金额
            amountRaw = str(row[7]).strip() if len(row) > 7 and row[7] else ""
            amount = None
            if amountRaw:
                try:
                    amount = float(amountRaw.replace(",", ""))
                except ValueError:
                    pass

            # 解析合格状态
            qualifiedRaw = str(row[8]).strip().upper() if len(row) > 8 and row[8] else ""
            qualified = qualifiedRaw == "YES"

            record = ProcessRecord(
                id=uuid.uuid4().hex,
                timestamp=str(row[1]) if row[1] else "",
                phone=str(row[2]) if row[2] else "",
                icNumber=str(row[3]) if row[3] else None,
                success=True,
                qualified=qualified,
                receiptNo=str(row[4]) if row[4] else None,
                brand=str(row[6]) if len(row) > 6 and row[6] else None,
                amount=amount,
                confidence=confidence,
                disqualifyReason=str(row[9]) if len(row) > 9 and row[9] else None,
                imagePath=str(row[11]) if len(row) > 11 and row[11] else None,
                rawText=None,
                error=None,
                steps=[],  # 历史记录无步骤详情
                totalDurationMs=0.0,
            )
            store.add(record)
            count += 1

        return count
    finally:
        wb.close()


# 全局单例
processStore = ProcessStore()
