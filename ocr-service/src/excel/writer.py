import asyncio
import os
from datetime import datetime
from typing import Optional
from concurrent.futures import ThreadPoolExecutor

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from ..config.loader import get_config
from .schema import (
    get_receipt_columns,
    get_receipts_sheet_name,
    get_registration_columns,
    get_registrations_sheet_name,
)

# 全局写入锁，防止并发写入同一 xlsx 文件
_write_lock = asyncio.Lock()
_executor = ThreadPoolExecutor(max_workers=1)  # 单线程池确保文件操作排队，且不阻塞主线程

# Excel 表头样式
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")


def _get_excel_path() -> str:
    """从配置读取 Excel 文件路径，转为绝对路径"""
    config = get_config()
    relative_path = config["excel"]["file_path"]
    # 相对路径基于项目根目录（ocr-service 的上两级）
    project_root = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "../../../")
    )
    return os.path.join(project_root, relative_path)


def _ensure_workbook(excel_path: str) -> Workbook:
    """
    加载已有工作簿或新建，并确保两个 Sheet 存在
    新建时自动添加表头行
    """
    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        os.makedirs(os.path.dirname(excel_path), exist_ok=True)
        wb = Workbook()
        # 删除默认的 Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    reg_sheet_name = get_registrations_sheet_name()
    rec_sheet_name = get_receipts_sheet_name()

    # 确保注册 Sheet 存在
    if reg_sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(reg_sheet_name)
        _write_header(ws, get_registration_columns())

    # 确保收据 Sheet 存在
    if rec_sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(rec_sheet_name)
        _write_header(ws, get_receipt_columns())

    return wb


def _write_header(ws, columns: list[str]) -> None:
    """写入带样式的表头行"""
    ws.append(columns)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")


def _next_seq(ws) -> int:
    """计算下一行序号（数据行数，不含表头）"""
    return ws.max_row  # 表头已占第1行，max_row 即序号


async def write_registration(phone: str, ic_number: str) -> bool:
    """
    追加写入注册记录到 Sheet1
    返回 True 表示写入成功
    """
    async with _write_lock:
        excel_path = _get_excel_path()
        
        def _do_write():
            wb = _ensure_workbook(excel_path)
            ws = wb[get_registrations_sheet_name()]
            seq = _next_seq(ws)
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.append([seq, now, phone, ic_number, "已注册"])
            wb.save(excel_path)
            return True
            
        loop = asyncio.get_event_loop()
        await loop.run_in_executor(_executor, _do_write)

    return True


async def write_receipt(
    phone: str,
    ic_number: str,
    receipt_no: Optional[str],
    raw_brand: Optional[str],
    matched_brand: Optional[str],
    amount: Optional[float],
    qualified: bool,
    disqualify_reason: Optional[str],
    confidence: float,
    image_path: Optional[str],
) -> int:
    """
    追加写入收据记录到 Sheet2
    返回写入的序号 seq
    """
    async with _write_lock:
        excel_path = _get_excel_path()
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        qualified_str = "YES" if qualified else "NO"
        amount_str = f"{amount:.2f}" if amount is not None else ""
        confidence_str = f"{confidence:.2%}"

        def _do_write():
            wb = _ensure_workbook(excel_path)
            ws = wb[get_receipts_sheet_name()]
            seq = _next_seq(ws)
            
            ws.append([
                seq,
                now,
                phone,
                ic_number,
                receipt_no or "",
                raw_brand or "",
                matched_brand or "",
                amount_str,
                qualified_str,
                disqualify_reason or "",
                confidence_str,
                image_path or "",
            ])
            wb.save(excel_path)
            return seq
            
        loop = asyncio.get_event_loop()
        seq = await loop.run_in_executor(_executor, _do_write)

    return seq


async def update_receipt_in_excel(
    seq: int,
    matched_brand: Optional[str],
    amount: Optional[float],
    qualified: bool,
    disqualify_reason: Optional[str],
    receipt_no: Optional[str]
) -> bool:
    """
    更新 Excel 中指定 seq 的收据记录
    """
    async with _write_lock:
        excel_path = _get_excel_path()
        if not os.path.exists(excel_path):
            return False
            
        def _do_write():
            wb = _ensure_workbook(excel_path)
            ws = wb[get_receipts_sheet_name()]

            row_index = seq + 1
            if row_index <= ws.max_row and ws.cell(row=row_index, column=1).value == seq:
                ws.cell(row=row_index, column=5).value = receipt_no or ""
                ws.cell(row=row_index, column=7).value = matched_brand or ""
                ws.cell(row=row_index, column=8).value = f"{amount:.2f}" if amount is not None else ""
                ws.cell(row=row_index, column=9).value = "YES" if qualified else "NO"
                ws.cell(row=row_index, column=10).value = disqualify_reason or ""
                wb.save(excel_path)
                return True
            return False
            
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(_executor, _do_write)


async def is_ic_registered(ic_number: str) -> bool:
    """检查身份证是否已注册（读操作，不需要写锁）"""
    excel_path = _get_excel_path()
    if not os.path.exists(excel_path):
        return False

    def _do_read():
        wb = load_workbook(excel_path, read_only=True, data_only=True)
        try:
            sheet_name = get_registrations_sheet_name()

            if sheet_name not in wb.sheetnames:
                return False

            ws = wb[sheet_name]
            # IC 在第4列（index 3），跳过表头行
            ic_col_index = 3
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ic_col_index] == ic_number:
                    return True

            return False
        finally:
            wb.close()
            
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(_executor, _do_read)
